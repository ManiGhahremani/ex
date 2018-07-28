#!/usr/bin/env python
import os
import openpyxl
import sys

#
#
#file1
#
#
#read in position of the first file
print("Hello, please make sure you run this file from its local directory.")
print("This script reads in a selection from the first file and pastes it in a selection in the second")
f1path = raw_input("Please enter the path to file1 relative to current folder for example ../exFolder/f1.xlsx: \n")

try:
    f1 = openpyxl.load_workbook(f1path)
except IOError:
    sys.exit('invalid path to file1, please re-try')

print("Now you have to select a sheet in file1:")
dec1 = raw_input("Enter 'P' to choose the position of the sheet, enter 'N' to choose the name of the sheet:\n")
if(dec1=='P'):
    dec2 = raw_input("Okey now type in the index of the sheet:\n")
    try:
        f1.active = int(dec2)
    except ValueError:
        sys.exit('invalid sheet index in file1, please re-try')
elif(dec1=='N'):
    dec2 = raw_input("Okey now type in the name of the sheet:\n")
    sheetlist1 = f1.sheetnames
    try:
        activeindex1 = sheetlist1.index(dec2)
    except ValueError:
        sys.exit('invalid sheet name in file1, please re-try')

sheet1=f1.active

#
#
#file2
#
#
f2path = raw_input("Now please enter the path to second file:\n")

try:
    f2 = openpyxl.load_workbook(f2path)
except IOError:
    sys.exit('invalid path to file2, please re-try')

print("Now you have to select a sheet in file2:")
dec3 = raw_input("Enter 'P' to choose the position of the sheet, enter 'N' to choose the name of the sheet:\n")
if(dec3=='P'):
    dec4 = raw_input("Okey now type in the index of the sheet:\n")
    try:
        f2.active = int(dec4)
    except ValueError:
        sys.exit('invalid sheet index in file2, please re-try')
elif(dec3=='N'):
    dec4 = raw_input("Okey now type in the name of the sheet:\n")
    sheetlist2 = f2.sheetnames
    try:
        activeindex2 = sheetlist2.index(dec4)
    except ValueError:
        sys.exit('invalid sheet name in file2, please re-try')

sheet2 = f2.active

#
#
#sel1
#
#
print("Now enter the selection of file1:")
c1f1 = raw_input("Please enter start column of file1 to copy as int:\n")
r1f1 = raw_input("Please enter start row of file1 to copy as int:\n")
c2f1 = raw_input("Please enter end column of file1 to copy as int:\n")
r2f1 = raw_input("Please enter end row of file1 to copy as int:\n")

print("Now enter the selection of file2 (Please make sure the selections are the same size):")
c1f2 = raw_input("Please enter start column of file2 to paste as int:\n")
r1f2 = raw_input("Please enter start row of file2 to paste as int:\n")
c2f2 = raw_input("Please enter end column of file2 to paste as int:\n")
r2f2 = raw_input("Please enter end row of file2 to paste as int:\n")

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

try:
    selectedRange = copyRange(c1f1, r1f1, c2f1, r2f1, sheet1)
except TypeError:
    sys.exit('the input selection was wrong type')

#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):

            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
try:
    pastingRange = pasteRange(c1f2, r1f2, c2f2, r2f2, sheet2, selectedRange)
except TypeError:
    sys.exit('the input selection was wrong type')

f2.save(filename=f2path)
print("Range copied and pasted!")
