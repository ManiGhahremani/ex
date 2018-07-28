#!/usr/bin/env python

import openpyxl
import os

# print the current working directory
print("\nThe current working directory is" + os.getcwd())
#change directory if needed
#os.chdir(../path/to/other/folder)
#or if needed make a new folder
#os.system('mkdir newfolder')
#similarly we can copy file
#os.system('copy t1.py ..\new/path/to/copy/file/t2.py')
#or move or delete or anything in the directory
#for more details read the cmd commands, and type
#them between the single quotes below:
#os.system('move t1.py ..\new/path/to/put/file/t1.py')
#we are making a backup
#os.system('cp e.xlsx ebackup.xlsx')


######################################################################################################


filename = 'e.xlsx'
#opening the file in the same folder
file = openpyxl.load_workbook(filename)


######################################################################################################

#getting sheets in a file
def printSheetTitles(file):
    print("\n")
    counter = 0
    for sheet in file:
        counter = counter + 1
        print("sheet number " + str(counter) + " is titled: " + sheet.title)

printSheetTitles(file)

#another way to get all the sheets' names
def printSheetAtPos(file, position):
    sheetlist = file.sheetnames
    selectedsheet = file[ sheetlist[position] ]
    print("\nThe selected sheet name is: " + str(selectedsheet.title))

printSheetAtPos(file,0)



######################################################################################################


#activate the sheet at position 0
file.active = 0
activesheet = file.active
#print out the name of the activated sheet
print("\nThe activated sheet name is: " + activesheet.title)
#or
#print(file.active.title)


######################################################################################################


#create a new sheet and put it at position 1
#(if there was anyother sheets at position 1, it would push them forwards.)
#sheet2 = file.create_sheet("SecondSheetTitile", 1)
#printSheetTitles(file)


######################################################################################################


def returnSheetAtPos(file, position):
    sheetlist = file.sheetnames
    sheetRet = file[ sheetlist[position] ]
    return sheetRet

#rename sheets:
sheet2 = returnSheetAtPos(file, 2)
sheet2.title='SecondSheetNewTitle'
file.save(filename)
#printSheetTitles(file)


######################################################################################################


#We can copy the active work sheet like this
#copy = file.copy_worksheet(returnSheetAtPos(file, 0))
#we must save changes for it to make effect


######################################################################################################


#create a new cell at a specific position
newcell = activesheet['H1']
#and we can change its value
activesheet['H1'] = '1'
print("\nH1 cell's row is: " + str(newcell.row) + " and column is: " + str(newcell.column)
      + " and its value is: " + str(newcell.value))
#and modify it again
activesheet['H1'] = '2'
#we must save before it takes effect


######################################################################################################


#for immediate change and access

activesheet.cell(row=2, column=2).value = 'U22222'
#we must save before it takes effect


######################################################################################################


#if we save it with the same filename, it basically saves changes
file.save(filename)
#we can create functions like this to help us code faster
#this function will save the file
def savecur():
    file.save(filename)
    return

#we use the function like this
savecur()
#but the problem is that we cannot use this function if we have many files


######################################################################################################


filename2 = 'e2.xlsx'
file2 = openpyxl.Workbook()
file2.active = 0

#so we can use this function instead
def savecur(n,f):
    f.save(filename = n)
    return

#and we call it like this
savecur('e2.xlsx', file2)


######################################################################################################


#we can change a range of cells like this easily
for row in range(1,5):
    for col in range(1,8):
        file2.active.cell(column=col, row=row).value = "c{0},r{1}".format(col,row)


######################################################################################################


file2.active = 0
f2sheetactive = file2.active
#We can get the number of rows and columns
rowCount = f2sheetactive.max_row
print("\n" + filename2 + " has " +str(rowCount) + " rows")
columnCount = f2sheetactive.max_column
print(filename2 + " has " + str(columnCount) + " columns")


######################################################################################################


#we want to select a range of cells in file2
f2cells = f2sheetactive['A1':'G2']
#this is saved as a tuple so lets save the size of our selection
f2cellsRCount = len(f2cells)
f2cellsCCount = 0
for f2cellsR in f2cells:
    if(f2cellsCCount<len(f2cellsR)):
        f2cellsCCount = len(f2cellsR)

print("\nThe tuple row number is:" + str(f2cellsRCount))
print("\nThe tuple column number is:" + str(f2cellsCCount))
#now we have the row count and col count of this tuple


######################################################################################################


#so we start by accessing tuple[0][0]]
#all the way up to tuple[f2cellsRCount][f2cellsCCount]
#so we have assigned the index i to rows, and j to columns
##everytime we access tuple[i][j] and we increment each accordingly
i=0
while(i<f2cellsRCount):
    for r in range(155,157):
        j=0
        while(j<f2cellsCCount):
            for c in range(1,8):
                activesheet.cell(column=c, row=r).value = f2cells[i][j].value
                j=j+1
        i=i+1


######################################################################################################

print("\nIterating rows")
#we can iterate through rows like this:
for row in file2.active.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in row:
        print cell.value

#same with columns
print("\nIterating columns")
#we can iterate through rows like this:
for col in file2.active.iter_cols(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in col:
        print cell.value


######################################################################################################


#we can store all rows of a file like this:
savedRows = tuple(file2.active.rows)
#same with columns
savedCols = tuple(file2.active.columns)

#we can use the function below to convert the tuples to lists if needed
def tuptupToLili(tuptup):
    lili = []
    for tup in tuptup:
        li = list(tup)
        lili.append(li)
    return lili


######################################################################################################


#for certain occasions maybe we need to import new data types
import datetime
#many different ways to do it
file2.active['D6'].value = datetime.datetime(2010, 7, 21)
file2.active['D6'].value = datetime.datetime.now()
print("\nDatetime value:")
print(file2.active['D6'].value)


######################################################################################################


#we can use formulas like This
file2.active['A6'].value = 5
a6 = file2.active['A6'].value
file2.active['B6'].value = 6
b6 = file2.active['B6'].value
#this is just sum of 4 and 5
file2.active['C6'].value = "=SUM(4, 5)"
#but this is sum of two other cells
file2.active['C7'].value = "=SUM({0}, {1})".format(a6, b6)
#so we can set the value of many cells like this, so they automatically change accordingly
#for more formulae look at "from openpyxl.utils import FORMULAE"



######################################################################################################


#we can save a file as a template like this:
#file2.template = True
#or choose to just save it as a file
file2.template = False
file2.save('e2.xlsx')


######################################################################################################

#we can merge a selection of cells like this
file2.active.merge_cells('A7:D8')
#to unmerge we can just do:
#file2.active.unmerge_cells('A7:D8')


######################################################################################################


#to work with images we need a new library
from openpyxl.drawing.image import Image

#this image exits in the directory, so between '' we put the path to file
img = Image('logo.png')
file2.active.add_image(img, 'A10')


######################################################################################################


#we can fold columns like this:
#(its called grouping)
file2.active.column_dimensions.group('H', 'J', hidden=True)


######################################################################################################



print("\n")
savecur(filename2, file2)
savecur(filename, file)
#os.system('rm e.xlsx')
#os.system('mv ebackup.xlsx e.xlsx')
